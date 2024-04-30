from openpyxl import Workbook, load_workbook

class Appointment:
    def __init__(self, patient_name, doctor_name, date, time):
        self.patient_name = patient_name
        self.doctor_name = doctor_name
        self.date = date
        self.time = time
        self.canceled = False  # Initialize the canceled attribute to False

class AppointmentScheduler:
    def __init__(self, filename):
        self.filename = filename
        self.appointments = {}
        self.load_appointments()

    def load_appointments(self):
        self.appointments = {}  # Clear the appointments dictionary before loading
        try:
            wb = load_workbook(self.filename)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                appointment = Appointment(*row)
                # Generate a unique identifier for each appointment
                appointment_key = f"{appointment.patient_name}_{appointment.doctor_name}_{appointment.date}_{appointment.time}"
                # Check if appointment is not canceled and not already present
                if not appointment.canceled and appointment_key not in self.appointments:
                    self.appointments[appointment_key] = appointment
            wb.close()
        except FileNotFoundError:
            print("File not found. Creating a new one.")
            self.create_excel()

    def create_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["Patient Name", "Doctor Name", "Date", "Time"])
        wb.save(self.filename)

    def schedule_appointment(self, appointment):
        # Generate a unique identifier for the new appointment
        appointment_key = f"{appointment.patient_name}_{appointment.doctor_name}_{appointment.date}_{appointment.time}"
        # Check if appointment is not canceled and not already present
        if not appointment.canceled and appointment_key not in self.appointments:
            self.appointments[appointment_key] = appointment
            self.update_excel()
            print("Appointment scheduled successfully!")
        else:
            print("Appointment could not be scheduled.")

    def update_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["Patient Name", "Doctor Name", "Date", "Time"])
        for appointment_key, appointment in self.appointments.items():
            ws.append([appointment.patient_name, appointment.doctor_name, appointment.date, appointment.time])
        wb.save(self.filename)
        wb.close()

    def view_appointments(self):
        if not self.appointments:
            print("No appointments scheduled.")
        else:
            print("Scheduled Appointments:")
            for idx, appointment in enumerate(self.appointments.values(), start=1):
                print(f"Appointment {idx}:")
                print(f"Patient: {appointment.patient_name}")
                print(f"Doctor: {appointment.doctor_name}")
                print(f"Date: {appointment.date}")
                print(f"Time: {appointment.time}")
                print()

    def cancel_appointment(self, patient_name, doctor_name, date, time):
        appointment_key = f"{patient_name}_{doctor_name}_{date}_{time}"
        if appointment_key in self.appointments:
            canceled_appointment = self.appointments.pop(appointment_key)
            print(f"Appointment with {canceled_appointment.patient_name} on {canceled_appointment.date} at {canceled_appointment.time} has been canceled.")
            self.update_excel()
        else:
            print("Appointment not found.")

def main():
    filename = "Appointment.xlsx"
    scheduler = AppointmentScheduler(filename)

    while True:
        print("1. Schedule an appointment")
        print("2. View appointments")
        print("3. Cancel an appointment")
        print("4. Exit")
        choice = input("Enter your choice: ")

        if choice == '1':
            patient_name = input("Enter patient's name: ")
            doctor_name = input("Enter doctor's name: ")
            date = input("Enter appointment date (YYYY-MM-DD): ")
            time = input("Enter appointment time (HH:MM): ")
            appointment = Appointment(patient_name, doctor_name, date, time)
            scheduler.schedule_appointment(appointment)
        elif choice == '2':
            scheduler.view_appointments()
        elif choice == '3':
            scheduler.view_appointments()
            if scheduler.appointments:
                patient_name = input("Enter patient's name: ")
                doctor_name = input("Enter doctor's name: ")
                date = input("Enter appointment date (YYYY-MM-DD): ")
                time = input("Enter appointment time (HH:MM): ")
                scheduler.cancel_appointment(patient_name, doctor_name, date, time)
        elif choice == '4':
            print("Exiting program...")
            break
        else:
            print("Invalid choice. Please try again.")

if __name__ == "__main__":
    main()
